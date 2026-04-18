import openpyxl
from collections import Counter

def generate_report(vehicle_log: list, duration: float, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws["A1"], ws["B1"] = "Total unique vehicles", len(vehicle_log)
    ws["A2"], ws["B2"] = "Processing duration (s)", duration

    ws["A4"], ws["B4"] = "Vehicle type", "Count"
    for i, (vtype, count) in enumerate(Counter(v["class"] for v in vehicle_log).items(), 5):
        ws[f"A{i}"], ws[f"B{i}"] = vtype, count

    ws2 = wb.create_sheet("Detections")
    ws2.append(["Tracker ID", "Class", "First Frame", "Timestamp (s)"])
    for v in vehicle_log:
        ws2.append([v["tracker_id"], v["class"], v["first_frame"], v["timestamp_s"]])

    wb.save(output_path)